from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .hashing import file_sha256, text_sha1
from .ids import make_seed_id


_HEADER_ROW = ("情境场景", "触发事件")


def _normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def _is_section_marker(text: str) -> bool:
    return text.startswith("【")


def _is_catalog_header_row(scene_text: str, trigger_text: str) -> bool:
    if trigger_text:
        return False
    normalized = scene_text.replace(" ", "")
    return bool(re.search(r"\bX_[A-Za-z]+\b", normalized)) or normalized in {
        "外部诱因空间(X_ext)",
        "内部诱因空间(X_int)",
    }


def _clean_text(text: str | None) -> str:
    return _normalize_spaces(str(text or "").replace("\n", " "))


def _trigger_core(trigger_text: str) -> str:
    trigger_text = re.sub(r"（.*?）", "", trigger_text)
    trigger_text = re.sub(r"\(.*?\)", "", trigger_text)
    trigger_text = trigger_text.split("：")[0].split(":")[0]
    return _normalize_spaces(trigger_text)


def build_query_seed_catalog(*, excel_path: str | Path, sheet_name: str) -> dict[str, Any]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows: list[dict[str, Any]] = []
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        scene_text = _clean_text(row[0] if len(row) > 0 else "")
        trigger_text = _clean_text(row[1] if len(row) > 1 else "")
        if not scene_text or scene_text == _HEADER_ROW[0]:
            continue
        if _is_catalog_header_row(scene_text, trigger_text):
            continue
        if _is_section_marker(scene_text) or _is_section_marker(trigger_text):
            continue
        seed_id = make_seed_id(index, scene_text, trigger_text)
        query_texts = [scene_text, f"{scene_text} 现场"]
        trigger_core = _trigger_core(trigger_text)
        if trigger_core and not _is_section_marker(trigger_core):
            composite = _normalize_spaces(f"{scene_text} {trigger_core}")
            if composite not in query_texts:
                query_texts.append(composite)
        rows.append(
            {
                "seed_id": seed_id,
                "excel_row": index,
                "scene_text": scene_text,
                "trigger_text": trigger_text,
                "query_texts": query_texts,
            }
        )
    return {
        "meta": {
            "source_path": str(Path(excel_path).resolve()),
            "source_sha256": file_sha256(excel_path),
            "sheet_name": sheet_name,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
        },
        "rows": rows,
    }


def _load_schema_rows(path: str | Path) -> list[dict[str, Any]]:
    source_path = Path(path)
    if source_path.suffix.lower() == ".json":
        return json.loads(source_path.read_text(encoding="utf-8"))
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    sheet = workbook.active
    raw_rows = list(sheet.iter_rows(values_only=True))
    headers = [str(cell or "").strip() for cell in raw_rows[0]]
    items: list[dict[str, Any]] = []
    for row in raw_rows[1:]:
        item = {}
        for index, header in enumerate(headers):
            item[header] = row[index] if index < len(row) else None
        items.append(item)
    return items


def _parse_allowed_values(raw_domain: str) -> list[str]:
    return [_normalize_spaces(item) for item in str(raw_domain).split("|") if _normalize_spaces(item)]


def build_annotation_domains(*, schema_source_path: str | Path, label_seed_path: str | Path) -> dict[str, Any]:
    schema_rows = _load_schema_rows(schema_source_path)
    label_seed = json.loads(Path(label_seed_path).read_text(encoding="utf-8"))
    label_values = list(dict.fromkeys(str(item).strip() for item in label_seed.get("labels", []) if str(item).strip()))

    fields: dict[str, Any] = {}
    for row in schema_rows:
        field_name = str(row.get("字段名") or "").strip()
        if not field_name:
            continue
        data_type = str(row.get("数据类型") or "").strip().lower()
        raw_domain = str(row.get("值域") or "").strip()
        required_text = str(row.get("必填") or "").strip()
        field_spec: dict[str, Any] = {
            "type": data_type or "string",
            "required": required_text in {"是", "type≠none 时必填"},
            "required_when": "trigger_event_type!=none" if required_text == "type≠none 时必填" else None,
            "description": str(row.get("说明") or "").strip(),
        }
        if field_name == "group_emotion":
            field_spec["allowed_values"] = label_values
        elif "1–9" in raw_domain or "1-9" in raw_domain:
            field_spec["type"] = "integer"
            field_spec["min_value"] = 1
            field_spec["max_value"] = 9
        elif data_type == "enum":
            field_spec["allowed_values"] = _parse_allowed_values(raw_domain)
        else:
            field_spec["allowed_values"] = None
        field_spec["allow_rejected_only_values"] = []
        if field_name == "group_size_estimate":
            field_spec["allow_rejected_only_values"] = ["single", "none"]
        elif field_name == "group_behavior_type":
            field_spec["allow_rejected_only_values"] = ["none"]
        elif field_name == "group_emotion":
            field_spec["allow_rejected_only_values"] = ["none"]
        fields[field_name] = field_spec

    return {
        "meta": {
            "source_schema_path": str(Path(schema_source_path).resolve()),
            "source_sha256": file_sha256(schema_source_path),
            "label_seed_path": str(Path(label_seed_path).resolve()),
            "label_seed_sha256": file_sha256(label_seed_path),
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "field_count": len(fields),
            "group_emotion_label_count": len(label_values),
        },
        "fields": fields,
    }


def prepare_reference_files(
    *,
    excel_path: str | Path,
    sheet_name: str,
    schema_source_path: str | Path,
    label_seed_path: str | Path,
    query_seed_catalog_path: str | Path,
    annotation_domains_path: str | Path,
) -> dict[str, Path]:
    query_seed_catalog = build_query_seed_catalog(excel_path=excel_path, sheet_name=sheet_name)
    annotation_domains = build_annotation_domains(schema_source_path=schema_source_path, label_seed_path=label_seed_path)

    query_seed_catalog_path = Path(query_seed_catalog_path)
    annotation_domains_path = Path(annotation_domains_path)
    query_seed_catalog_path.parent.mkdir(parents=True, exist_ok=True)
    annotation_domains_path.parent.mkdir(parents=True, exist_ok=True)

    query_seed_catalog_path.write_text(json.dumps(query_seed_catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    annotation_domains_path.write_text(json.dumps(annotation_domains, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "query_seed_catalog_path": query_seed_catalog_path,
        "annotation_domains_path": annotation_domains_path,
    }
